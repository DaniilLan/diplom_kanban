from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.shortcuts import redirect, render
from django.utils.formats import localize
from django.db.models import Q
from .forms import NewUserForm
from .models import Group, Task, TimeLog
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from .models import UsersGroup
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.db.models import Sum
from django.utils import timezone
from .models import Task, TimeLog, BoardNames, TaskType, PriorityTask



@login_required
def home(request):
    user_groups = Group.objects.filter(members__user=request.user)
    tasks = Task.objects.filter(
        Q(owner=request.user) |
        Q(groups__in=user_groups)
    ).distinct()

    all_tasks = []
    for t in tasks:
        # Получаем список групп задачи
        task_groups = t.groups.all()

        t_dict = {
            'uuid': str(t.uuid),
            'name': t.name if t.name is not None else 'Без названия',
            'boardName': t.boardName,
            'date': str(localize(t.date)),
            'task_id': str(t.task_id),
            'description': str(t.description),
            'typeTask': str(t.typeTask),
            'priorityTask': str(t.priorityTask),
            'timeEstimateMinutes': str(t.timeEstimateMinutes),
            'groups': [
                {
                    'uuid': str(g.uuid),
                    'name': g.name
                }
                for g in task_groups
            ],
            'responsible': str(t.responsible),
        }
        all_tasks.append(t_dict)

    all_timelogs = []
    accessible_timelogs = TimeLog.objects.filter(task__in=tasks)
    for tl in accessible_timelogs:
        t_dict = {
            "task": str(tl.task.uuid),
            "minutesSpent": tl.minutesSpent,
            "date": tl.date.isoformat(),
            "comment": tl.comment,
            "owner_username": tl.owner.username
        }
        all_timelogs.append(t_dict)
    return render(request, 'index.html', {'tasks': all_tasks, 'timelogs': all_timelogs})

def register_request(request):
    if request.method == 'POST':
        form = NewUserForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request,
                             'Аккаунт зарегистрирован: '
                             'добро пожаловать на сайт!')
            return redirect('board:login')
        messages.error(request, 'Не удалось зарегистрировать аккаунт. '
                                'Проверьте корректность данных и '
                                'попробуйте еще раз!')
    form = NewUserForm()
    return render(request=request,
                  template_name='register.html',
                  context={'register_form': form})


def login_request(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.info(request,
                              f'Вы вошли на сайт под ником {username}.')
                return redirect('board:home')
            else:
                messages.error(request, 'Неверные имя и/или пароль.')
        else:
            messages.error(request, 'Неверные имя и/или пароль.')
    form = AuthenticationForm()
    return render(request=request, template_name='login.html',
                  context={'login_form': form})


def logout_request(request):
    logout(request)
    messages.info(request, 'Вы вышли из аккаунта.')
    return redirect('board:login')


def export_to_excel(request):
    # Создаем новую книгу Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Kanban Доска"

    # Заголовки столбцов
    headers = [
        "ID", "Название", "Описание", "Статус",
        "Тип", "Приоритет", "Владелец", "Ответственный",
        "Дата создания", "Оценка времени", "Затраченное время", "Группы"
    ]

    # Добавляем заголовки
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

    # Получаем все задачи пользователя
    tasks = Task.objects.filter(owner=request.user).order_by('boardName')

    # Добавляем данные
    for row_num, task in enumerate(tasks, 2):
        # ID задачи
        task_type_prefix = 'BUG-' if task.typeTask == TaskType.BUG else 'TASK-'
        ws.cell(row=row_num, column=1, value=f"{task_type_prefix}{task.task_id}")

        # Основные данные
        ws.cell(row=row_num, column=2, value=task.name)
        ws.cell(row=row_num, column=3, value=task.description or "")

        # Статус - преобразуем через BoardNames
        status_mapping = {
            BoardNames.ToDo: 'Сделать',
            BoardNames.ReOpen: 'Переоткрыто',
            BoardNames.InProgress: 'В работе',
            BoardNames.Review: 'На проверке',
            BoardNames.InTest: 'В тестировании',
            BoardNames.Done: 'Выполнено'
        }
        ws.cell(row=row_num, column=4, value=status_mapping.get(task.boardName, task.boardName))

        # Тип задачи - используем get_typeTask_display()
        ws.cell(row=row_num, column=5, value=task.get_typeTask_display())

        # Приоритет - используем get_priorityTask_display()
        ws.cell(row=row_num, column=6, value=task.get_priorityTask_display())

        # Владелец и ответственный
        ws.cell(row=row_num, column=7, value=task.owner.username)
        responsible = task.responsible.username if task.responsible else ""
        ws.cell(row=row_num, column=8, value=responsible)

        # Дата создания
        ws.cell(row=row_num, column=9, value=task.date.strftime('%Y-%m-%d %H:%M'))

        # Оценка времени
        estimate_time = convert_minutes_to_readable(task.timeEstimateMinutes)
        ws.cell(row=row_num, column=10, value=estimate_time)

        # Затраченное время
        total_minutes = TimeLog.objects.filter(task=task).aggregate(total=Sum('minutesSpent'))['total'] or 0
        spent_time = convert_minutes_to_readable(total_minutes)
        ws.cell(row=row_num, column=11, value=spent_time)

        # Группы
        groups = ", ".join([group.name for group in task.groups.all()])
        ws.cell(row=row_num, column=12, value=groups)

    # Настраиваем ширину столбцов
    column_widths = {
        'A': 15,  # ID
        'B': 40,  # Название
        'C': 60,  # Описание
        'D': 15,  # Статус
        'E': 10,  # Тип
        'F': 12,  # Приоритет
        'G': 15,  # Владелец
        'H': 15,  # Ответственный
        'I': 20,  # Дата создания
        'J': 15,  # Оценка времени
        'K': 15,  # Затраченное время
        'L': 30  # Группы
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Добавляем границы и выравнивание
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            if cell.column_letter in ['C', 'L']:  # Описание и группы
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            else:
                cell.alignment = Alignment(vertical='center')

    # Фиксируем заголовки
    ws.freeze_panes = 'A2'

    # Создаем HTTP-ответ с файлом Excel
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=kanban_board_{timezone.now().strftime("%Y-%m-%d")}.xlsx'
    wb.save(response)

    return response


def convert_minutes_to_readable(minutes):
    """Конвертирует минуты в читаемый формат (дни, часы, минуты)"""
    if not minutes:
        return "0м"

    days = minutes // (8 * 60)  # Предполагаем 8-часовой рабочий день
    remaining = minutes % (8 * 60)
    hours = remaining // 60
    mins = remaining % 60

    parts = []
    if days > 0:
        parts.append(f"{days}д")
    if hours > 0:
        parts.append(f"{hours}ч")
    if mins > 0 or not parts:
        parts.append(f"{mins}м")

    return " ".join(parts)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_group_members(request):
    user = request.user
    # Получаем все группы пользователя
    user_groups = UsersGroup.objects.filter(user=user).values_list('group', flat=True)
    # Получаем всех участников этих групп
    members = UsersGroup.objects.filter(group__in=user_groups).select_related('user')
    # Собираем уникальные username
    unique_usernames = set(member.user.username for member in members)
    return Response(list(unique_usernames))


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_task_group_members(request, task_uuid):
    try:
        task = Task.objects.get(uuid=task_uuid)
        # Получаем все группы задачи
        groups = task.groups.all()
        # Получаем всех участников этих групп
        members = UsersGroup.objects.filter(group__in=groups).select_related('user')
        # Собираем уникальные username
        unique_usernames = set(member.user.username for member in members)
        return Response(list(unique_usernames))
    except Task.DoesNotExist:
        return Response({'error': 'Task not found'}, status=404)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_user_groups(request):
    user = request.user
    groups = UsersGroup.objects.filter(user=user).select_related('group')
    groups_data = [{
        'uuid': ug.group.uuid,
        'name': ug.group.name
    } for ug in groups]
    return Response(groups_data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def get_task_groups(request, task_uuid):
    try:
        task = Task.objects.get(uuid=task_uuid)
        groups = task.groups.all().values('uuid', 'name')
        return Response(groups)
    except Task.DoesNotExist:
        return Response({'error': 'Task not found'}, status=404)